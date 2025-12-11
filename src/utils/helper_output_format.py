from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, numbers



def style_excel_header(file_path, sheet_name=None):
  
    wb = load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    normal_font = Font(bold=False)
    left_align = Alignment(horizontal='left')
    no_border = Border(
        left=Side(border_style=None),
        right=Side(border_style=None),
        top=Side(border_style=None),
        bottom=Side(border_style=None)
    )

    for cell in ws[1]:
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = no_border

    

    wb.save(file_path)
    print(f'[Styler] Header formatting applied → {file_path}')
